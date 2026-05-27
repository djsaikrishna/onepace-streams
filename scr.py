import os
import re
import json
import time
import random
import requests
import openpyxl
import datetime
from bs4 import BeautifulSoup

SHEET_EXPORT_URL = "https://docs.google.com/spreadsheets/d/1HQRMJgu_zArp-sLnvFMDzOyjdsht87eFLECxMK858lA/export?format=xlsx"
LOCAL_EXCEL_FILE = "one_pace.xlsx"
TRACKER_FILE = "tracker.json"

# --- Load Central Config ---
try:
    with open('config.json', 'r', encoding='utf-8') as f:
        CONFIG = json.load(f)
except (FileNotFoundError, json.JSONDecodeError) as e:
    print(f"CRITICAL ERROR: Failed to load config.json. {e}")
    exit(1)
PREFIX_MAP = CONFIG["ARC_MAP"]
ALIASES = CONFIG.get("ALIASES", {})

# Global caches to prevent spamming Nyaa
resolved_batches_cache = {}
nyaa_html_cache = {}
WEBSITE_HTML_CACHE = None

def download_excel_file(url, filename, max_retries=3):
    print(f"Downloading latest spreadsheet to {filename}...")
    for attempt in range(1, max_retries + 1):
        try:
            response = requests.get(url, stream=True, timeout=60)
            response.raise_for_status()
            with open(filename, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk: f.write(chunk)
            print("Download complete!\n")
            return True
        except Exception as e:
            print(f"  [!] Attempt {attempt}/{max_retries} failed: {e}")
            if attempt < max_retries:
                print("  [*] Retrying in 3 seconds...")
                time.sleep(3)
            else:
                print("  [-] All attempts to download the spreadsheet failed.")
                return False

def resolve_nyaa_url(query, max_retries=2):
    rss_url = f"https://nyaa.si/?page=rss&q={query}&c=0_0&f=0"
    for attempt in range(1, max_retries + 1):
        try:
            response = requests.get(rss_url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=15)
            response.raise_for_status()
            match = re.search(r'<guid isPermaLink="true">(https://nyaa\.si/view/\d+)</guid>', response.text)
            if match:
                return match.group(1)
            break 
        except requests.exceptions.RequestException:
            if attempt < max_retries: time.sleep(random.uniform(1, 3))

    html_url = f"https://nyaa.si/?q={query}&f=0&c=0_0"
    for attempt in range(1, max_retries + 1):
        try:
            response = requests.get(html_url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=15)
            response.raise_for_status()
            match = re.search(r'href="(/view/\d+)"', response.text)
            if match:
                return f"https://nyaa.si{match.group(1)}"
            return None
        except requests.exceptions.RequestException:
            if attempt < max_retries: time.sleep(random.uniform(1, 3))
    return None

def resolve_nyaa_batch(arc_name, max_retries=3):
    search_name = ALIASES.get(arc_name, arc_name)
    clean_name = search_name.replace("The Adventures of ", "").replace("The Trials of ", "").strip()
    
    query = f"One Pace {clean_name}".replace("'", "").replace(' ', '+')
    html_url = f"https://nyaa.si/?f=0&c=0_0&q={query}"
    
    for attempt in range(1, max_retries + 1):
        try:
            response = requests.get(html_url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=15)
            response.raise_for_status()
            matches = re.findall(r'<a href="(/view/\d+)" title="([^"]+)">', response.text)
            
            best_link, best_score = None, -1
            
            for link, title in matches:
                if "One Pace" not in title or clean_name not in title or link.endswith("#comments"):
                    continue
                
                score = 0
                if "1080" in title: score += 30
                elif "720" in title: score += 20
                elif "480" in title: score += 10
                
                is_batch = False
                if "Batch" in title or "batch" in title.lower():
                    is_batch = True
                elif re.search(rf'{clean_name}\s+\d{{1,3}}\s*[-~]\s*\d{{1,3}}', title, re.IGNORECASE):
                    is_batch = True
                elif not re.search(rf'{clean_name}\s+\d{{1,3}}\b', title, re.IGNORECASE):
                    is_batch = True
                    
                if is_batch: score += 100 
                    
                if score > best_score:
                    best_score = score
                    best_link = f"https://nyaa.si{link}"

            if best_link and best_score >= 100:
                print(f"  [✅] Found Best Batch URL (Score {best_score}): {best_link}")
                return best_link
            return None
        except requests.exceptions.RequestException:
            if attempt < max_retries: time.sleep(random.uniform(2, 4))
    return None

def resolve_single_episode(arc_name, ep_num, max_retries=2):
    search_name = ALIASES.get(arc_name, arc_name)
    clean_name = search_name.replace("The Adventures of ", "").replace("The Trials of ", "").strip()
    
    ep_padded = str(ep_num).zfill(2)
    query = f"One Pace {clean_name} {ep_padded}".replace("'", "").replace(' ', '+')
    html_url = f"https://nyaa.si/?f=0&c=0_0&q={query}"
    
    for attempt in range(1, max_retries + 1):
        try:
            response = requests.get(html_url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=15)
            response.raise_for_status()
            matches = re.findall(r'<a href="(/view/\d+)" title="([^"]+)">', response.text)
            
            best_link, best_score = None, -1
            for link, title in matches:
                if "One Pace" not in title or clean_name not in title or link.endswith("#comments"):
                    continue
                if not re.search(rf'\b{ep_padded}\b|\b{ep_num}\b', title):
                    continue
                    
                score = 30 if "1080" in title else 20 if "720" in title else 10 if "480" in title else 0
                if score > best_score:
                    best_score = score
                    best_link = f"https://nyaa.si{link}"

            if best_link:
                print(f"  [✅] Found Single Episode URL: {best_link}")
                return best_link
            return None
        except requests.exceptions.RequestException:
            if attempt < max_retries: time.sleep(random.uniform(1, 3))
    return None

def get_torrent_data(nyaa_url, expected_ep_num, expected_crc=None, max_retries=3):
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
    html_content = None
    
    if nyaa_url in nyaa_html_cache:
        html_content = nyaa_html_cache[nyaa_url]
    else:
        for attempt in range(1, max_retries + 1):
            try:
                response = requests.get(nyaa_url, headers=headers, timeout=15)
                response.raise_for_status()
                html_content = response.text
                nyaa_html_cache[nyaa_url] = html_content
                break
            except requests.exceptions.RequestException as e:
                print(f"  [!] Nyaa timeout on attempt {attempt}/{max_retries}. Error: {e}")
                if attempt < max_retries:
                    time.sleep(random.uniform(3, 6))
                else:
                    return None, None, None # <-- Updated return

    if not html_content:
        return None, None, None # <-- Updated return

    soup = BeautifulSoup(html_content, 'html.parser')
    info_hash = None
    torrent_filename = "Unknown Title"
    file_idx = None # <-- NEW: Initialize as None

    magnet_tag = soup.find('a', href=re.compile(r'^magnet:\?xt=urn:btih:'))
    if magnet_tag:
        match = re.search(r'urn:btih:([a-zA-Z0-9]{40})', magnet_tag['href'])
        if match:
            info_hash = match.group(1).lower()

    title_tag = soup.find('title')
    if title_tag:
        raw_title = title_tag.text
        torrent_filename = raw_title.replace(" :: Nyaa", "").strip()

    file_list_div = soup.find('div', class_=re.compile('torrent-file-list'))
    if file_list_div:
        ep_padded = str(expected_ep_num).zfill(2)
        lines = file_list_div.get_text(separator='\n').split('\n')
        
        video_files = []
        for line in lines:
            text = line.strip()
            if ".mkv" in text or ".mp4" in text:
                clean_file = re.sub(r'\s*\([^)]*\)$', '', text).strip()
                video_files.append(clean_file)
        
        if video_files:
            video_files.sort()
            matched = False
            
            # --- STRICT CRC MATCHING ---
            if expected_crc and expected_crc not in ["Unknown", "00000000"]:
                for idx, vf in enumerate(video_files):
                    if expected_crc.upper() in vf.upper() or expected_crc.lower() in vf.lower():
                        torrent_filename = vf
                        # Set to None if it's the only file, otherwise save the index
                        file_idx = None if len(video_files) == 1 else idx 
                        matched = True
                        break
                        
                if not matched:
                    return None, None, None # <-- Updated return
            
            # --- FALLBACK: Episode Number Matching (Only if no CRC provided) ---
            if not matched:
                if len(video_files) == 1:
                    torrent_filename = video_files[0]
                    file_idx = None # <-- Null for single files
                    matched = True
                else:
                    for idx, vf in enumerate(video_files):
                        if re.search(rf'\b{ep_padded}\b|\b{expected_ep_num}\b', vf):
                            torrent_filename = vf
                            file_idx = idx # <-- Save the index
                            matched = True
                            break
                            
            # --- FINAL FALLBACK: Index ---
            if not matched:
                ep_index = int(expected_ep_num) - 1
                if 0 <= ep_index < len(video_files):
                    torrent_filename = video_files[ep_index]
                    file_idx = ep_index # <-- Save the index
                else:
                    return None, None, None # <-- Updated return

    if info_hash:
        return info_hash, torrent_filename, file_idx # <-- Updated return
    return None, None, None # <-- Updated return

# --- NEW: Global cache to track previous episode numbers ---
LAST_EPISODE_CACHE = {}

def get_expected_filename(ep_name, arc_name):
    prefix = PREFIX_MAP.get(arc_name, arc_name[:2].upper())
    ep_name_str = str(ep_name).strip()

    # Initialize the tracker for this arc if it doesn't exist yet
    if arc_name not in LAST_EPISODE_CACHE:
        LAST_EPISODE_CACHE[arc_name] = 0

    # --- Custom Override for Alternate / G-8 Episodes ---
    if "Alternate" in ep_name_str or "(G8)" in ep_name_str:
        match = re.search(r'\b(\d{1,3})\b', ep_name_str)
        if match:
            ep_num = int(match.group(1)) + 1
        else:
            # Fallback: Last known episode + 1
            ep_num = LAST_EPISODE_CACHE[arc_name] + 1
            
        LAST_EPISODE_CACHE[arc_name] = ep_num
        return f"{prefix}_{ep_num}.json"

    # --- Standard extraction for regular episodes ---
    match = re.search(r'(\d+)\s*$', ep_name_str)
    if not match: 
        match = re.search(r'\b(\d{1,3})\b', ep_name_str)
        
    if match:
        ep_num = int(match.group(1))
    else:
        # Fallback: Last known episode + 1
        ep_num = LAST_EPISODE_CACHE[arc_name] + 1
    
    # Save this episode number as the new "last seen" for this arc
    LAST_EPISODE_CACHE[arc_name] = ep_num
    
    return f"{prefix}_{ep_num}.json"

def save_tracker(tracker_data):
    with open(TRACKER_FILE, 'w', encoding='utf-8') as f:
        json.dump(tracker_data, f, indent=2, ensure_ascii=False)

def resolve_from_website(arc_name, ep_num_raw, max_retries=2):   
    global WEBSITE_HTML_CACHE
    url = "https://onepace.net/en/releases"
    search_name = ALIASES.get(arc_name, arc_name)
    clean_name = search_name.replace("The Adventures of ", "").replace("The Trials of ", "").strip().lower()
    
    if not WEBSITE_HTML_CACHE:
        for attempt in range(1, max_retries + 1):
            try:
                response = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=15)
                response.raise_for_status()
                WEBSITE_HTML_CACHE = BeautifulSoup(response.text, 'html.parser')
                break
            except requests.exceptions.RequestException:
                if attempt < max_retries: time.sleep(random.uniform(1, 3))
                
    if not WEBSITE_HTML_CACHE: return [] # Return an empty list instead of None

    episodes = WEBSITE_HTML_CACHE.find_all('li', id=True) 
    found_streams = [] # Store all cuts here!
    
    for ep in episodes:
        title_tag = ep.find('h3')
        if not title_tag: continue
        
        title_text = title_tag.get_text(separator=" ", strip=True).lower()
        if clean_name not in title_text: continue
        
        # --- NEW: Ignore Archived Episodes! ---
        if "archived" in title_text: continue
        
        # Strict Episode Matching (Fixes the "5 days ago" bug)
        match = re.search(rf'{re.escape(clean_name)}\s*(\d{{1,3}})(?:\s*-\s*(\d{{1,3}}))?', title_text)
        if not match: continue
        
        ep_start = int(match.group(1))
        ep_end = int(match.group(2)) if match.group(2) else ep_start
        
        if not (ep_start <= int(ep_num_raw) <= ep_end):
            continue
        
        time_tag = ep.find('time')
        if not time_tag or not time_tag.has_attr('datetime'): continue
        
        try:
            release_date = datetime.datetime.strptime(time_tag['datetime'][:10], "%Y-%m-%d").date()
            days_diff = (datetime.datetime.now().date() - release_date).days
            if days_diff > 7 or days_diff < 0: continue 
        except: continue
        
        # --- Check the <small> tag for "Extended" ---
        small_tag = title_tag.find('small')
        is_extended = small_tag and "extended" in small_tag.text.lower()
        
        magnet_tag = ep.find('a', href=re.compile(r'^magnet:\?xt='))
        if magnet_tag:
            match = re.search(r'urn:btih:([a-zA-Z0-9]{40})', magnet_tag['href'], re.IGNORECASE)
            if match:
                view_url = resolve_nyaa_url(match.group(1).lower())
                if view_url: 
                    found_streams.append({"url": view_url, "is_extended": is_extended})
                    continue # Successfully found magnet, move to next ep block
        
        torrent_tag = ep.find('a', href=re.compile(r'nyaa\.si/download/\d+\.torrent'))
        if torrent_tag:
            match = re.search(r'/download/(\d+)\.torrent', torrent_tag['href'])
            if match: 
                found_streams.append({"url": f"https://nyaa.si/view/{match.group(1)}", "is_extended": is_extended})
                
    return found_streams

def main():
    start_time = time.time()
    new_files_count = 0 
    updated_files_list = [] 
    
    output_dir = "stream"
    os.makedirs(output_dir, exist_ok=True)
    
    tracker_data = {}
    if os.path.exists(TRACKER_FILE):
        with open(TRACKER_FILE, 'r', encoding='utf-8') as f:
            tracker_data = json.load(f)
            for key, val in tracker_data.items():
                if isinstance(val, str):
                    tracker_data[key] = [val]
            
    success = download_excel_file(SHEET_EXPORT_URL, LOCAL_EXCEL_FILE)
    if not success: return

    print("Loading local spreadsheet...")
    workbook = openpyxl.load_workbook(LOCAL_EXCEL_FILE)
    
    files_to_process = {}
    episode_lengths = {} 
    
    for target_sheet in workbook.sheetnames:
        if target_sheet not in PREFIX_MAP:
            continue
            
        sheet = workbook[target_sheet]
        print(f"\n--- Scanning Arc: {target_sheet} ---")
        
        ep_col_idx, header_row = None, None
        length_col_indices = [] 
        
        for row in range(1, 10):
            for col in range(1, sheet.max_column + 1):
                cell_val = str(sheet.cell(row=row, column=col).value).strip()
                if "One Pace Episode" in cell_val: 
                    ep_col_idx = col
                    header_row = row
                elif "Length" in cell_val: 
                    length_col_indices.append(col)           
        if not ep_col_idx:
            print(f"  [!] Could not find the Episode Name column. Skipping.")
            continue

        for row in range(header_row + 1, sheet.max_row + 1):
            ep_name = sheet.cell(row=row, column=ep_col_idx).value
            if not ep_name: continue
            
            filename = get_expected_filename(ep_name, target_sheet)
            
            row_lengths = []
            for l_col in length_col_indices:
                val = sheet.cell(row=row, column=l_col).value
                if val:
                    if isinstance(val, datetime.time) or isinstance(val, datetime.datetime):
                        row_lengths.append(val.strftime("%M:%S"))
                    else:
                        row_lengths.append(str(val).strip())

            if filename not in files_to_process:
                files_to_process[filename] = {"urls": [], "arc": target_sheet} # Store arc name here
                episode_lengths[filename] = {}
            
            row_urls = []
            for col in range(1, sheet.max_column + 1):
                if col == ep_col_idx: continue 
                cell = sheet.cell(row=row, column=col)
                target = None
                
                if cell.hyperlink and cell.hyperlink.target:
                    target = cell.hyperlink.target
                elif cell.value and isinstance(cell.value, str) and 'HYPERLINK' in cell.value:
                    match = re.search(r'HYPERLINK\("([^"]+)"', cell.value)
                    if match: target = match.group(1)
                elif cell.value and isinstance(cell.value, str):
                    val = cell.value.strip()
                    if re.match(r'^[A-Fa-f0-9]{8}$', val):
                        target = f"BATCH_SEARCH:{target_sheet}:{val}"
                    elif val.isdigit() and 6 <= len(val) <= 8:
                        target = f"https://nyaa.si/view/{val}"
                    # --- NEW: Catch raw 40-character infohashes just in case! ---
                    elif re.match(r'^[A-Fa-f0-9]{40}$', val):
                        target = f"https://nyaa.si/?q={val}"
                        
                if target and ("nyaa.si" in target or "magnet:" in target or target.startswith("BATCH_SEARCH:")):
                    if target not in row_urls:
                        row_urls.append(target)

            # --- SMART FAILSAFE: Prevent guessing unreleased episodes ---
            if not row_urls:
                # Unreleased episodes have no video length. 
                # If it HAS a length but no link, it's a typo (like SK_14), so we hunt for it!
                has_valid_length = any(l and str(l).strip() and "TBA" not in str(l).upper() for l in row_lengths)
                if has_valid_length:
                    row_urls.append(f"BATCH_SEARCH:{target_sheet}:Unknown")

            for idx, url in enumerate(row_urls):
                if url not in files_to_process[filename]["urls"]:
                    files_to_process[filename]["urls"].append(url)
                    assigned_length = row_lengths[idx] if idx < len(row_lengths) else ""
                    episode_lengths[filename][url] = assigned_length

    print("\n--- Processing Streams & Saving JSONs ---")
    
    # --- REPLACE THE START OF THE PROCESSING LOOP ---
    for filename, info in files_to_process.items():
        nyaa_urls = info["urls"]
        arc_name = info["arc"] # Correctly retrieve the arc for this specific file
        
        if not nyaa_urls: continue
            
        filepath = os.path.join(output_dir, filename)
        ep_num_raw = filename.split('_')[-1].replace('.json', '')
        
        # --- PRE-CHECK: DOES THE WEBSITE HAVE A FRESH RELEASE? ---
        # Call it ONCE and save the list to website_streams
        website_streams = resolve_from_website(arc_name, ep_num_raw)
        
        # Skip ONLY if tracker matches AND there is NO fresh website release overriding it
        if not website_streams and tracker_data.get(filename) == nyaa_urls and os.path.exists(filepath):
            print(f"  [~] Skipped {filename} (Already up-to-date)")
            continue
            
        print(f"  [*] Processing {filename} (Found {len(nyaa_urls)} stream(s)!)")
        
        streams = []

        # Extract lengths from the spreadsheet (Standard is usually index 0, Extended is index 1)
        spreadsheet_lengths = list(episode_lengths.get(filename, {}).values())

        # --- STREAM SLOTS (Prevents Old vs New Duplicates) ---
        has_standard = False
        has_extended = False

        # --- PHASE 0: OFFICIAL WEBSITE OVERRIDE ---
        for web_stream in website_streams:
            web_info_hash, web_filename, web_file_idx = get_torrent_data(web_stream["url"], ep_num_raw, expected_crc=None)
            if web_info_hash:
                is_ext = web_stream["is_extended"]
                ext_label = "Extended" if is_ext else "Standard"
                print(f"  [⭐] Found fresh {ext_label} release on Website: {web_filename}")
                
                assigned_len = ""
                if is_ext and len(spreadsheet_lengths) > 1:
                    assigned_len = spreadsheet_lengths[1]
                elif not is_ext and len(spreadsheet_lengths) > 0:
                    assigned_len = spreadsheet_lengths[0]

                streams.append({
                    "infoHash": web_info_hash, 
                    "filename": web_filename, 
                    "length": assigned_len,
                    "fileIdx": web_file_idx  # <-- Added here!
                })
                
                # Lock the slot!
                if is_ext: has_extended = True
                else: has_standard = True

        # --- PHASES 1-3: SPREADSHEET CRCs & EXTENDED CUTS ---
        for idx, url in enumerate(nyaa_urls):
            # idx 0 = Standard Slot, idx 1 = Extended Slot
            is_extended_col = (idx == 1)

            # If the website already filled the slot we are currently looking at, skip!
            if is_extended_col and has_extended:
                print("  [⚡] Skipping outdated Spreadsheet Extended Cut (already got fresh one).")
                continue
            if not is_extended_col and has_standard:
                print("  [⚡] Skipping outdated Spreadsheet Standard Cut (already got fresh one).")
                continue

            info_hash, torrent_filename, file_idx = None, None, None # <-- Add file_idx
            actual_url = url
            expected_crc = None
            
            if actual_url.startswith("BATCH_SEARCH:"):
                parts = actual_url.split(":")
                expected_crc = parts[2] if len(parts) > 2 else None

            # --- PHASE 1: Handle raw InfoHash queries from the spreadsheet ---
            if "nyaa.si/?q=" in actual_url:
                match = re.search(r'\?q=([a-fA-F0-9]{40})', actual_url)
                if match:
                    hash_val = match.group(1)
                    print(f"  [🔎] Resolving Spreadsheet InfoHash: {hash_val}...")
                    resolved = resolve_nyaa_url(hash_val)
                    actual_url = resolved if resolved else f"BATCH_SEARCH:{arc_name}:Unknown"
            
            # --- PHASE 2: Test Direct URLs (If it's not a batch search) ---
            if not actual_url.startswith("BATCH_SEARCH:"):
                info_hash, torrent_filename, file_idx = get_torrent_data(actual_url, ep_num_raw, expected_crc=None)
                if not info_hash:
                    print(f"  [⏳] Direct link failed. Triggering Fallback...")
                    actual_url = f"BATCH_SEARCH:{arc_name}:Unknown"
            
            # --- PHASE 3: The Ultimate Fallback Chain ---
            if actual_url.startswith("BATCH_SEARCH:"):
                parts = actual_url.split(":")
                arc_name = parts[1]
                expected_crc = parts[2] if len(parts) > 2 else None
                
                # TIER 1: ALWAYS TRY THE OFFICIAL ARC BATCH FIRST
                if not info_hash:
                    batch_url = resolved_batches_cache.get(arc_name)
                    if not batch_url:
                        batch_url = resolve_nyaa_batch(arc_name)
                        if batch_url:
                            resolved_batches_cache[arc_name] = batch_url
                            time.sleep(random.uniform(1, 2))
                    
                    if batch_url:
                        info_hash, torrent_filename, file_idx = get_torrent_data(batch_url, ep_num_raw, expected_crc)
                
                # TIER 2: TRY PRECISE CRC SEARCH (If missing from batch)
                if not info_hash and expected_crc and expected_crc not in ["Unknown", "00000000"]:
                    print(f"  [🔎] Not in batch. Resolving CRC directly: {expected_crc}...")
                    crc_url = resolve_nyaa_url(expected_crc)
                    if crc_url:
                        info_hash, torrent_filename, file_idx = get_torrent_data(crc_url, ep_num_raw, expected_crc)
                        if info_hash: time.sleep(random.uniform(1, 2))
                            
                # TIER 3: SINGLE EPISODE SEARCH (Final Fallback)
                if not info_hash:
                    print(f"  [⏳] CRC missed. Hunting for Single Episode: {arc_name} {ep_num_raw}...")
                    sing_url = resolve_single_episode(arc_name, ep_num_raw)
                    if sing_url:
                        info_hash, torrent_filename, file_idx = get_torrent_data(sing_url, ep_num_raw, None)
                        if info_hash: time.sleep(random.uniform(1, 2))
                
                if not info_hash:
                    print(f"  [-] ❌ Could not resolve torrent for CRC, Episode, or Batch: {arc_name}")
                    continue
                
            if info_hash:
                # Slot is open! Add it.
                url_length = episode_lengths.get(filename, {}).get(url, "")
                streams.append({
                    "infoHash": info_hash, 
                    "filename": torrent_filename, 
                    "length": url_length,
                    "fileIdx": file_idx # <-- Added here!
                })
                
                # Lock the slot!
                if is_extended_col: has_extended = True
                else: has_standard = True
                
                if actual_url not in nyaa_html_cache:
                    time.sleep(random.uniform(0.5, 1.5))
        
        if streams:
            data = {"streams": streams}
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            
            tracker_data[filename] = nyaa_urls
            save_tracker(tracker_data)
            
            print(f"  [+] Saved {filename}")
            new_files_count += 1
            updated_files_list.append(filename) 
            
        else:
            print(f"  [-] Failed to get any infoHashes for {filename}")

    if updated_files_list:
        with open("stream/st_purge.txt", "w") as f:
            for fname in updated_files_list:
                f.write(fname + "\n")

    end_time = time.time()
    total_time = round(end_time - start_time, 2)
    
    print("\n=========================================")
    print(" ✅ SCRIPT FINISHED ")
    print("=========================================")
    print(f" ⏱️  Time taken: {total_time} seconds")
    print(f" 📥 New files downloaded: {new_files_count}")
    print("=========================================\n")

if __name__ == "__main__":
    main()